from logging import Logger
from pathlib import Path
from unittest.mock import Mock

import openpyxl as oxl
import pytest

from tesliper import Energies
from tesliper.extraction import Soxhlet
from tesliper.glassware import Conformers, SingleSpectrum, Spectra
from tesliper.writing.xlsx_writer import XlsxWriter


@pytest.fixture
def filenames():
    return ["meoh-1.out", "meoh-2.out"]


@pytest.fixture
def fixturesdir():
    return Path(__file__).parent.with_name("fixtures")


@pytest.fixture
def mols(filenames, fixturesdir):
    s = Soxhlet(fixturesdir)
    s.wanted_files = filenames
    return Conformers(s.extract())


@pytest.fixture
def spc(filenames):
    return SingleSpectrum(
        "ir",
        [0.3, 0.2, 10, 300, 2],
        [10, 20, 30, 40, 50],
        width=5,
        fitting="gaussian",
        filenames=filenames,
        averaged_by="gib",
    )


@pytest.fixture
def spectra(filenames):
    return Spectra(
        genre="ir",
        values=[[0.3, 0.2, 10, 300, 2], [0.5, 0.8, 12, 150, 5]],
        abscissa=[10, 20, 30, 40, 50],
        width=5,
        fitting="gaussian",
        filenames=filenames,
    )


@pytest.fixture
def writer(tmp_path):
    return XlsxWriter(tmp_path)


def test_start_with_existing_file(tmp_path):
    file = tmp_path.joinpath("tesliper-output.xlsx")
    wb = oxl.Workbook()
    wb.save(file)
    XlsxWriter(tmp_path, mode="a")


def test_energies(writer, mols):
    grn = "scf"
    writer.energies(mols.arrayed(grn))
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    assert wb.sheetnames == [f"distribution-{grn}"]
    ws = wb[f"distribution-{grn}"]
    assert len(list(ws.columns)) == 5
    assert len(list(ws.rows)) == 1 + len(list(mols.keys()))


def test_overview(writer, mols):
    writer.overview(
        [mols.arrayed(grn) for grn in Energies.associated_genres],
        frequencies=mols.arrayed("freq"),
        stoichiometry=mols.arrayed("stoichiometry"),
    )
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    assert wb.sheetnames == ["overview"]
    ws = wb["overview"]
    assert len(list(ws.columns)) == 13
    assert len(list(ws.rows)) == 2 + len(list(mols.keys()))


def test_energies_with_corrections(writer, mols):
    grn = "gib"
    corrs = mols.arrayed(f"{grn}corr")
    ens = mols.arrayed(grn)
    writer.energies(ens, corrections=corrs)
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    assert wb.sheetnames == [f"distribution-{grn}"]
    ws = wb[f"distribution-{grn}"]
    assert len(list(ws.columns)) == 6
    assert len(list(ws.rows)) == 1 + len(list(mols.keys()))


def test_activities(writer, mols, filenames):
    writer.spectral_activities(mols.arrayed("freq"), [mols.arrayed("iri")])
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    keys = [f"{Path(f).stem}.activities-vibrational" for f in filenames]
    assert wb.sheetnames == keys
    for file in keys:
        ws = wb[file]
        assert len(list(ws.columns)) == 2
        assert len(list(ws.rows)) == 1 + mols.arrayed("freq").values.shape[1]


def test_spectra(writer, mols, spectra):
    writer.spectra(spectra)
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    ws = wb[spectra.genre]
    assert len(list(ws.columns)) == 1 + spectra.filenames.size
    assert len(list(ws.rows)) == 1 + spectra.values.shape[1]


def test_single_spectrum(writer, mols, spc):
    writer.single_spectrum(spc)
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    ws = wb[f"spectrum.{spc.genre}-{spc.averaged_by}"]
    assert len(list(ws.columns)) == 2
    assert len(list(ws.rows)) == 1 + spc.values.size


@pytest.fixture
def filenamestd():
    return ["fal-td.out"]


@pytest.fixture
def molstd(filenamestd, fixturesdir):
    s = Soxhlet(fixturesdir)
    s.wanted_files = filenamestd
    return Conformers(s.extract())


def test_transitions_only_highest(writer, molstd, filenamestd):
    trans, wave = molstd.arrayed("transitions"), molstd.arrayed("wavelen")
    writer.transitions(trans, wave, only_highest=True)
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    keys = [f"{Path(f).stem}.transitions-highest" for f in filenamestd]
    assert wb.sheetnames == keys
    for file in keys:
        ws = wb[file]
        assert len(list(ws.columns)) == 5
        assert len(list(ws.rows)) == 1 + trans.values.shape[1]


def test_transitions_all(writer, molstd, filenamestd):
    trans, wave = molstd.arrayed("transitions"), molstd.arrayed("wavelen")
    writer.transitions(trans, wave, only_highest=False)
    assert writer.file.exists()
    wb = oxl.load_workbook(writer.file)
    keys = [f"{Path(f).stem}.transitions-all" for f in filenamestd]
    assert wb.sheetnames == keys
    for num, file in enumerate(keys):
        ws = wb[file]
        assert len(list(ws.columns)) == 5
        assert len(list(ws.rows)) == 1 + trans.values[num].count()


def test_not_implemented_write(writer, arrays, monkeypatch):
    monkeypatch.setattr(Logger, "warning", Mock())
    writer.write(arrays)
    #  Geometry and generic InfoArray not supported
    assert Logger.warning.call_count == 2

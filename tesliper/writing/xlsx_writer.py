# IMPORTS
import logging as lgg
from itertools import chain, repeat, zip_longest
from pathlib import Path
from string import Template
from typing import Iterable, Optional, Sequence, Union

import numpy as np
import openpyxl as oxl

from ..glassware.arrays import (
    Bands,
    DataArray,
    Energies,
    FloatArray,
    InfoArray,
    SpectralActivities,
    SpectralData,
    Transitions,
)
from ..glassware.spectra import SingleSpectrum, Spectra
from ._writer import Writer

# LOGGER
logger = lgg.getLogger(__name__)
logger.setLevel(lgg.DEBUG)


# CLASSES
class XlsxWriter(Writer):
    """Writes extracted data to .xlsx file.

    Parameters
    ----------
    destination : str or pathlib.Path
        Directory, to which generated files should be written.
    mode : str
        Specifies how writing to file should be handled. Should be one of characters:
         'a' (append to existing file), 'x' (only write if file doesn't exist yet),
         or 'w' (overwrite file if it already exists).
    filename : str or string.Template
        Filename of created .xlsx file or a template for generation of the name using
        `Writer.make_name` method.
    """

    extension = "xlsx"

    def __init__(
        self,
        destination: Union[str, Path],
        mode: str = "x",
        filename: str = "tesliper-output.${ext}",
    ):
        super().__init__(destination=destination, mode=mode)
        file = self.destination / Template(filename).substitute(ext=self.extension)
        self.file = self.check_file(file)
        if self.mode == "a":
            self.workbook = oxl.load_workbook(self.file)
        else:
            self.workbook = oxl.Workbook()
            self.workbook.remove(self.workbook.active)

    def overview(
        self,
        energies: Sequence[Energies],
        frequencies: Optional[DataArray] = None,
        stoichiometry: Optional[InfoArray] = None,
        name_template: Union[str, Template] = "${cat}",
    ):
        """Writes summarized information from multiple Energies objects to xlsx file.
        Creates a worksheet with energy values and calculated
        populations for each energy object given, as well as number of imaginary
        frequencies and stoichiometry of conformers if `frequencies` and `stoichiometry`
        are provided, respectively.

        Parameters
        ----------
        energies: list of glassware.Energies
            Energies objects that are to be exported
        frequencies: glassware.DataArray, optional
            DataArray object containing frequencies
        stoichiometry: glassware.InfoArray, optional
            InfoArray object containing stoichiometry information
        name_template : str or string.Template
            Template that will be used to generate filenames,
            defaults to "${cat}".
        """
        wb = self.workbook
        template_params = {"cat": "overview", "conf": "multiple"}
        ws = wb.create_sheet(title=self.make_name(name_template, **template_params))
        ens_no = len(energies)
        headers = ["Gaussian output file", "Populations / %", "Energies / hartree"]
        headers += ["Imag"] if frequencies is not None else []
        headers += ["Stoichiometry"] if stoichiometry is not None else []
        cells = [
            "A1",
            "B1",
            f"{chr(66+ens_no)}1",
            f"{chr(66+2*ens_no)}1",
            f"{chr(67+2*ens_no)}1",
        ]
        for header, cell in zip(headers, cells):
            ws[cell] = header
        names = [self._header[en.genre] for en in energies]
        ws.append([""] + names + names)
        ws.merge_cells("A1:A2")
        ws.merge_cells(f"B1:{chr(65+ens_no)}1")
        ws.merge_cells(f"{chr(66+ens_no)}1:{chr(65+2*ens_no)}1")
        if frequencies is not None or stoichiometry is not None:
            ws.merge_cells(f"{chr(66+2*ens_no)}1:{chr(66+2*ens_no)}2")
        if frequencies is not None and stoichiometry is not None:
            ws.merge_cells(f"{chr(67+2*ens_no)}1:{chr(67+2*ens_no)}2")
        ws.freeze_panes = "A3"
        filenames = energies[0].filenames
        fmts = (
            ["0"]
            + ["0.00%"] * len(energies)
            + ["0." + "0" * (8 if en.genre == "scf" else 6) for en in energies]
            + ["0", "0"]
        )
        values = [en.values for en in energies]
        populs = [en.populations for en in energies]
        imag = frequencies.imaginary if frequencies is not None else []
        stoich = stoichiometry.values if stoichiometry is not None else []
        rows = zip_longest(filenames, *populs, *values, imag, stoich)
        for row_num, values in enumerate(rows):
            filtered_values = ((f, v) for f, v in zip(fmts, values) if v is not None)
            for col_num, (fmt, value) in enumerate(filtered_values):
                cell = ws.cell(row=row_num + 3, column=col_num + 1)
                cell.value = value
                cell.number_format = fmt
        # set cells width
        widths = [0] + [10] * ens_no + [16] * ens_no
        widths += [6] if frequencies is not None else []
        widths += [0] if stoichiometry is not None else []
        for column, width in zip(ws.columns, widths):
            if not width:
                width = max(len(str(cell.value)) for cell in column) + 2
            column_letter = oxl.utils.get_column_letter(column[0].column)
            ws.column_dimensions[column_letter].width = width
        wb.save(self.file)
        logger.info("Overview export to xlsx files done.")

    def energies(
        self,
        energies: Energies,
        corrections: Optional[FloatArray] = None,
        name_template: Union[str, Template] = "distribution-${genre}",
    ):
        """Writes detailed information from multiple Energies objects to xlsx file.
        Creates one worksheet for each Energies  object provided.
        The sheet contains energy values, energy difference to lowest-energy conformer,
        Boltzmann factor, population of each conformer and corrections,
        if those are provided.

        Parameters
        ----------
        energies: list of glassware.Energies
            Energies objects that are to be exported
        corrections: list of glassware.DataArray
            DataArray objects containing energies corrections
        name_template : str or string.Template
            Template that will be used to generate filenames,
            defaults to "distribution-${genre}".
        """
        wb = self.workbook
        fmts = (
            ["0", "0.00%"]
            + ["0.0000"] * 2
            + ["0.00000000" if energies.genre == "scf" else "0.000000"] * 2
        )
        template_params = {
            "conf": "multiple",
            "genre": energies.genre,
            "cat": "populations",
        }
        ws = wb.create_sheet(title=self.make_name(name_template, **template_params))
        ws.freeze_panes = "A2"
        header = [
            "Gaussian output file",
            "Population / %",
            "Min. B. Factor",
            "DE / (kcal/mol)",
            "Energy / Hartree",
        ]
        header += ["Correction / Hartree"] if corrections is not None else []
        ws.append(header)
        corr_values = corrections.values if corrections is not None else []
        rows = zip_longest(
            energies.filenames,
            energies.populations,
            energies.min_factors,
            energies.deltas,
            energies.values,
            corr_values,
        )
        for row_num, values in enumerate(rows):
            filtered_values = ((f, v) for f, v in zip(fmts, values) if v is not None)
            for col_num, (fmt, value) in enumerate(filtered_values):
                cell = ws.cell(row=row_num + 2, column=col_num + 1)
                cell.value = value
                cell.number_format = fmt
        # set cells width
        widths = [0, 15, 14, 15, 16, 19]
        for column, width in zip(ws.columns, widths):
            if not width:
                width = max(len(str(cell.value)) for cell in column) + 2
            column_letter = oxl.utils.get_column_letter(column[0].column)
            ws.column_dimensions[column_letter].width = width
        wb.save(self.file)
        logger.info("Energies export to xlsx files done.")

    def spectral_data(
        self,
        band: SpectralActivities,
        data: Iterable[SpectralData],
        name_template: Union[str, Template] = "${conf}.${cat}-${det}",
    ):
        """Writes SpectralData objects to xlsx file (one sheet for each conformer).

        Parameters
        ----------
        band: glassware.SpectralActivities
            object containing information about band at which transitions occur;
            it should be frequencies for vibrational data and wavelengths or
            excitation energies for electronic data
        data: iterable of glassware.SpectralData
            SpectralData objects that are to be serialized; all should contain
            information for the same conformers. Assumes that all `data`'s elements have
            the same `spectra_type`, which is passed to the `name_template` as "det".
        name_template : str or string.Template
            Template that will be used to generate filenames,
            defaults to "${conf}.${cat}-${det}".

        Raises
        ------
        ValueError
            if `data` is an empty sequence
        """
        self._spectral(
            band=band, data=data, name_template=name_template, category="data"
        )

    def spectral_activities(
        self,
        band: SpectralActivities,
        data: Iterable[SpectralActivities],
        name_template: Union[str, Template] = "${conf}.${cat}-${det}",
    ):
        """Writes SpectralActivities objects to xlsx file (one sheet for each conformer).

        Parameters
        ----------
        band: glassware.SpectralActivities
            object containing information about band at which transitions occur;
            it should be frequencies for vibrational data and wavelengths or
            excitation energies for electronic data
        data: iterable of glassware.SpectralActivities
            SpectralActivities objects that are to be serialized; all should contain
            information for the same conformers. Assumes that all `data`'s elements have
            the same `spectra_type`, which is passed to the `name_template` as "det".
        name_template : str or string.Template
            Template that will be used to generate filenames,
            defaults to "${conf}.${cat}-${det}".

        Raises
        ------
        ValueError
            if `data` is an empty sequence
        """
        self._spectral(
            band=band, data=data, name_template=name_template, category="activities"
        )

    def _spectral(
        self,
        band: SpectralActivities,
        data: Union[Iterable[SpectralData], Iterable[SpectralActivities]],
        name_template: Union[str, Template],
        category: str,
    ):
        """Writes SpectralActivities objects to xlsx file (one sheet for each conformer).

        Parameters
        ----------
        band: glassware.SpectralActivities
            object containing information about band at which transitions occur;
            it should be frequencies for vibrational data and wavelengths or
            excitation energies for electronic data
        data: iterable of glassware.SpectralActivities or iterable of SpectralData
            SpectralActivities or SpectralData objects that are to be serialized; all
            should contain information for the same conformers. Assumes that all
            `data`'s elements have the same `spectra_type`, which is passed to the
            `name_template` as "det".
        name_template : str or string.Template
            Template that will be used to generate filenames.
        category : str
            Category of exported data genres.

        Raises
        ------
        ValueError
            if `data` is an empty sequence
        """
        wb = self.workbook
        data = [band] + list(data)
        try:
            spectra_type = data[1].spectra_type
        except IndexError:
            raise ValueError("No data to export.")
        genres = [bar.genre for bar in data]
        headers = [self._header[genre] for genre in genres]
        widths = [max(len(h), 10) for h in headers]
        fmts = [self._excel_formats[genre] for genre in genres]
        values = list(zip(*[bar.values for bar in data]))
        template_params = {"genre": band.genre, "cat": category, "det": spectra_type}
        for num, (fname, values_) in enumerate(zip(data[0].filenames, values)):
            template_params.update({"conf": fname, "num": num})
            ws = wb.create_sheet(title=self.make_name(name_template, **template_params))
            ws.append(headers)
            ws.freeze_panes = "B2"
            for column, width in zip(ws.columns, widths):
                column_letter = oxl.utils.get_column_letter(column[0].column)
                ws.column_dimensions[column_letter].width = width
            for col_num, (vals, fmt) in enumerate(zip(values_, fmts)):
                for row_num, v in enumerate(vals):
                    cell = ws.cell(row=row_num + 2, column=col_num + 1)
                    cell.value = v
                    cell.number_format = fmt
        wb.save(self.file)
        logger.info("SpectralActivities export to xlsx files done.")

    def spectra(
        self,
        spectra: Spectra,
        name_template: Union[str, Template] = "${genre}",
    ):
        """Writes given spectral data collectively to one sheet of xlsx workbook.

        Parameters
        ----------
        spectra: glassware.Spectra
            Spectra object, that is to be serialized
        name_template : str or string.Template
            Template that will be used to generate filenames,
            defaults to "${genre}".
        """
        wb = self.workbook
        template_params = {"genre": spectra.genre, "cat": "spectra"}
        ws = wb.create_sheet(title=self.make_name(name_template, **template_params))
        ws.freeze_panes = "B2"
        A0 = spectra.units["x"]
        ws.append([A0] + list(spectra.filenames))
        title = (
            f"{spectra.genre} calculated with peak width = "
            f'{spectra.width} {spectra.units["width"]} and '
            f"{spectra.fitting} fitting, shown as "
            f'{spectra.units["x"]} vs. {spectra.units["y"]}'
        )
        ws["A1"].comment = oxl.comments.Comment(title, "Tesliper")
        for line in zip(spectra.x, *spectra.y):
            ws.append(line)
        wb.save(self.file)
        logger.info("Spectra export to xlsx file done.")

    def single_spectrum(
        self,
        spectrum: SingleSpectrum,
        name_template: Union[str, Template] = "${cat}.${genre}-${det}",
    ):
        """Writes SingleSpectrum object to new sheet of xlsx workbook.

        Parameters
        ----------
        spectrum: glassware.SingleSpectrum
            spectrum, that is to be serialized
        name_template : str or string.Template
            Template that will be used to generate sheet names,
            defaults to "${cat}.${genre}-${det}".
        """
        # TODO: add comment as in txt export
        wb = self.workbook
        template_params = {
            "genre": spectrum.genre,
            "cat": "spectrum",
            "det": spectrum.averaged_by,
        }
        ws = wb.create_sheet(title=self.make_name(name_template, **template_params))
        ws.append([spectrum.units["x"], spectrum.units["y"]])
        for row in zip(spectrum.x, spectrum.y):
            ws.append(row)
        wb.save(self.file)
        logger.info("Spectrum export to xlsx files done.")

    def transitions(
        self,
        transitions: Transitions,
        wavelengths: Bands,
        only_highest=True,
        name_template: Union[str, Template] = "${conf}.${cat}-${det}",
    ):
        """Writes electronic transitions data to xlsx file
        (one sheet for each conformer).

        Parameters
        ----------
        transitions : glassware.Transitions
            Electronic transitions data that should be serialized.
        wavelengths : glassware.ElectronicActivities
            Object containing information about wavelength at which transitions occur.
        only_highest : bool
            Specifies if only transition of highest contribution to given band should
            be reported. If `False` all transition are saved to file.
            Defaults to `True`.
        name_template : str or string.Template
            Template that will be used to generate filenames,
            defaults to "${conf}.${cat}-${det}".
        """
        transtions_data = (
            transitions.highest_contribution
            if only_highest
            else (
                transitions.ground,
                transitions.excited,
                transitions.values,
                transitions.contribution,
            )
        )
        wb = self.workbook
        headers = [
            self._header[wavelengths.genre],
            "Ground",
            "Excited",
            "Coefficient",
            "Contribution",
        ]
        widths = [len(h) for h in headers]
        fmts = [self._excel_formats[wavelengths.genre], "0", "0", "0.0000", "0%"]

        template_params = {
            "genre": transitions.genre,
            "cat": "transitions",
            "det": "highest" if only_highest else "all",
        }
        for num, (fname, grounds, exciteds, values, contribs, bands) in enumerate(
            zip(
                transitions.filenames,
                *transtions_data,
                wavelengths.wavelen,
            )
        ):
            template_params.update({"conf": fname, "num": num})
            ws = wb.create_sheet(title=self.make_name(name_template, **template_params))
            ws.append(headers)
            ws.freeze_panes = "B2"
            for column, width in zip(ws.columns, widths):
                column_letter = oxl.utils.get_column_letter(column[0].column)
                ws.column_dimensions[column_letter].width = width
            row_num = 1
            for g, e, v, c, b in zip(grounds, exciteds, values, contribs, bands):
                try:
                    values_ = [
                        # print wavelength value only once
                        d
                        for d in zip(chain([b], repeat(None)), g, e, v, c)
                        # omit entry if any value is masked
                        if all(x is not np.ma.masked for x in d)
                    ]
                except TypeError:
                    # transition_data is transitions.highest_contribution
                    values_ = [(b, g, e, v, c)]
                for vals in values_:
                    row_num += 1
                    for col_num, (v_, fmt) in enumerate(zip(vals, fmts), start=1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = v_
                        cell.number_format = fmt
        wb.save(self.file)
        logger.info("Transitions export to xlsx files done.")

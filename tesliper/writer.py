# IMPORTS
import csv
import numpy as np
import logging as lgg
import os
import openpyxl as oxl
from itertools import zip_longest
from collections import OrderedDict

from . import glassware as gw


# LOGGER
logger = lgg.getLogger(__name__)
logger.setLevel(lgg.DEBUG)


# CLASSES
class Writer:

    writers = dict()

    def __init_subclass__(cls, fmt='', **kwargs):
        if not fmt:
            raise TypeError('Required keyword argument "fmt" not found.')
        if not hasattr(cls, 'write'):
            raise AttributeError(
                'Class derived from Writer should provide write method.'
            )
        super().__init_subclass__(**kwargs)
        Writer.writers[fmt] = cls
        logger.info(f'Writer {cls} registered for export to {fmt} file format.')

    _header = dict(
        freq='Frequencies',
        mass='Red. masses',
        frc='Frc consts',
        raman='Raman Activ',
        depolarp=r'Depolar \(P\)',
        depolaru=r'Depolar \(U\)',
        ramact='RamAct',
        depp='Dep-P',
        depu='Dep-U',
        alpha2='Alpha2',
        beta2='Beta2',
        alphag='AlphaG',
        gamma2='Gamma2',
        delta2='Delta2',
        cid1='CID1',
        raman2='Raman2',
        roa2='ROA2',
        cid2='CID2',
        raman3='Raman3',
        roa3='ROA3',
        cid3='CID3',
        rc180='RC180',
        rot='Rot. Str.',
        dip='Dip. Str.',
        roa1='ROA1',
        raman1='Raman1',
        ex_en='Excit. Energy',
        wave='Wavelenght',
        vrot='Rot.(velo)',
        lrot='Rot. (len)',
        vosc='Osc.(velo)',
        losc='Osc. (len)',
        iri='IR Int.',
        emang='E-M Angle',
        eemang='E-M Angle',
        zpe='Zero-point',
        ten='Thermal',
        ent='Enthalpy',
        gib='Gibbs',
        scf='SCF'
    )

    _formatters = dict(
        rot='{:> 10.4f}',
        dip='{:> 10.4f}',
        roa1='{:> 10.4f}',
        raman1='{:> 10.4f}',
        vrot='{:> 10.4f}',
        lrot='{:> 10.4f}',
        vosc='{:> 10.4f}',
        losc='{:> 10.4f}',
        iri='{:> 10.4f}',
        emang='{:> 10.4f}',
        eemang='{:> 10.4f}',
        zpe='{:> 13.4f}',
        ten='{:> 13.4f}',
        ent='{:> 13.4f}',
        gib='{:> 13.4f}',
        scf='{:> 13.4f}',
        ex_en='{:> 13.4f}',
        freq='{:> 10.2f}',
        wave='{:> 10.2f}',
        mass='{:> 11.4f}',
        frc='{:> 10.4f}',
        raman='{:> 11.4f}',
        depolarp='{:> 11.4f}',
        depolaru='{:> 11.4f}',
        ramact='{:> 10.4f}',
        depp='{:> 9.4f}',
        depu='{:> 9.4f}',
        alpha2='{:> 9.4f}',
        beta2='{:> 9.4f}',
        alphag='{:> 9.4f}',
        gamma2='{:> 9.4f}',
        delta2='{:> 9.4f}',
        cid1='{:> 8.3f}',
        raman2='{:> 8.3f}',
        roa2='{:> 8.3f}',
        cid2='{:> 8.3f}',
        raman3='{:> 8.3f}',
        roa3='{:> 8.3f}',
        cid3='{:> 8.3f}',
        rc180='{:> 8.3f}',
    )

    _excel_formats = dict(
        freq='0.0000',
        mass='0.0000',
        frc='0.0000',
        raman='0.0000',
        depolarp='0.0000',
        depolaru='0.0000',
        ramact='0.0000',
        depp='0.0000',
        depu='0.0000',
        alpha2='0.0000',
        beta2='0.0000',
        alphag='0.0000',
        gamma2='0.0000',
        delta2='0.0000',
        cid1='0.000',
        raman2='0.000',
        roa2='0.000',
        cid2='0.000',
        raman3='0.000',
        roa3='0.000',
        cid3='0.000',
        rc180='0.000',
        rot='0.0000',
        dip='0.0000',
        roa1='0.000',
        raman1='0.000',
        ex_en='0.0000',
        wave='0.0000',
        vrot='0.0000',
        lrot='0.0000',
        vosc='0.0000',
        losc='0.0000',
        iri='0.0000',
        emang='0.0000',
        eemang='0.0000',
        zpe='0.000000',
        ten='0.000000',
        ent='0.000000',
        gib='0.000000',
        scf='0.00000000'
    )

    energies_order = 'zpe ten ent gib scf'.split(' ')

    def distribute_data(self, data):
        distr = dict(
            energies=[],
            vibra=[],
            electr=[],
            other_bars=[],
            spectra=[],
            single=[],
            other=[],
            corrections={},
            frequencies=None,
            wavelenghts=None,
            stoichiometry=None
        )
        for obj in data:
            if isinstance(obj, gw.Energies):
                distr['energies'].append(obj)
            elif obj.genre.endswith('corr'):
                distr['corrections'][obj.genre[:3]] = obj
            elif obj.genre == 'freq':
                distr['frequencies'] = obj
            elif obj.genre == 'wave':
                distr['wavelengths'] = obj
            elif obj.genre == 'stoichiometry':
                distr['stoichiometry'] = obj
            elif isinstance(obj, gw.Bars):
                if obj.spectra_type == 'vibra':
                    distr['vibra'].append(obj)
                elif obj.spectra_type == 'electr':
                    distr['electr'].append(obj)
                else:
                    distr['other_bars'].append(obj)
            elif isinstance(obj, gw.SingleSpectrum):
                distr['single'].append(obj)
            elif isinstance(obj, gw.Spectra):
                distr['spectra'].append(obj)
            else:
                distr['other'].append(obj)
        return distr


class TxtWriter(Writer, fmt='txt'):

    def write(self, dest, data):
        data = self.distribute_data(data)
        if data['energies']:
            file = os.path.join(dest, 'distribution_overview.txt')
            self.energies_overview(
                file, data['energies'], frequencies=data['frequencies'],
                stoichiometry=data['stoichiometry']
            )
            for ens in data['energies']:
                file = os.path.join(dest, f'distribution.{ens.genre}.txt')
                self.energies(
                    file, ens, corrections=data['corrections'].get(ens.genre)
                )
        if data['vibra']:
            self.bars(
                dest, band=data['frequencies'], bars=data['vibra'],
                interfix='vibra'
            )
        if data['electr']:
            self.bars(
                dest, band=data['wavelengths'], bars=data['electr'],
                interfix='electr'
            )
        if data['other_bars']:
            # TO DO
            pass
        if data['spectra']:
            for spc in data['spectra']:
                self.spectra(dest, spc, interfix=spc.genre)
        if data['single']:
            for spc in data['single']:
                interfix = f'.{spc.averaged_by}' if spc.averaged_by else ''
                file = os.path.join(dest, f'spectrum.{spc.genre+interfix}.txt')
                self.single_spectrum(file, spc)
        if data['other']:
            # TO DO
            pass

    def energies(self, file, energies, corrections=None):
        """Writes Energies object to txt file.

        Parameters
        ----------
        file: string
            path to file
        energies: glassware.Energies
            Energies object that is to be serialized
        corrections: glassware.DataArray, optional
            DataArray object, containing energies corrections"""
        max_fnm = max(np.vectorize(len)(energies.filenames).max(), 20)
        # file_path = os.path.join(self.path,
        #                          f'distribution.{energies.genre}.txt')
        header = [f"{'Gaussian output file':<{max_fnm}}"]
        header += ['Population/%', 'Min.B.Factor', 'DE/(kcal/mol)',
                   'Energy/Hartree']
        header += ['Corr/Hartree'] if corrections is not None else []
        header = ' | '.join(header)
        align = ('<', '>', '>', '>', '>', '>')
        width = (max_fnm, 12, 12, 13, 14, 12)
        corrections = corrections.values if corrections is not None else []
        fmt = ('', '.4f', '.4f', '.4f',
               '.8f' if energies.genre == 'scf' else '.6f', 'f')
        rows = zip_longest(energies.filenames, energies.populations * 100,
                           energies.min_factors, energies.deltas,
                           energies.values, corrections, fillvalue=None)
        with open(file, 'w') as file_:
            file_.write(header + '\n')
            file_.write('-' * len(header) + '\n')
            for row in rows:
                new_row = [f'{v:{a}{w}{f}}'
                           for v, a, w, f in zip(row, align, width, fmt)
                           if v is not None]
                file_.write(' | '.join(new_row) + '\n')
        logger.info('Energies separate export to text files done.')

    def energies_overview(self, file, energies, frequencies=None,
                          stoichiometry=None):
        """Writes essential information from multiple Energies objects to
         single txt file.

        Parameters
        ----------
        file: string
            path to file
        energies: list of glassware.Energies
            Energies objects that is to be expored
        frequencies: glassware.DataArray, optional
            DataArray object containing frequencies
        stoichiometry: glassware.InfoArray, optional
            InfoArray object containing stoichiometry information"""
        filenames = energies[0].filenames
        imaginary = [] if frequencies is None else frequencies.imaginary
        stoichiometry = [] if stoichiometry is None else stoichiometry.values
        max_fnm = max(np.vectorize(len)(filenames).max(), 20)
        try:
            max_stoich = max(np.vectorize(len)(stoichiometry).max(), 13)
        except ValueError:
            max_stoich = 0
        values = np.array([en.values for en in energies]).T
        # deltas = np.array([en.deltas for en in ens])
        popul = np.array([en.populations * 100 for en in energies]).T
        _stoich = f" | {'Stoichiometry':<{max_stoich}}"
        names = [self._header[en.genre] for en in energies]
        population_widths = [max(8, len(n)) for n in names]
        population_subheader = '  '.join(
            [f'{n:<{w}}' for n, w in zip(names, population_widths)]
        )
        energies_widths = [14 if n == 'SCF' else 12 for n in names]
        energies_subheader = '  '.join(
            [f'{n:<{w}}' for n, w in zip(names, energies_widths)]
        )
        precisions = [8 if n == 'SCF' else 6 for n in names]
        header = f"{'Gaussian output file':<{max_fnm}} | " \
                 f"{'Population / %':^{len(population_subheader)}} | " \
                 f"{'Energy / Hartree':^{len(energies_subheader)}}" \
                 f"{' | Imag' if frequencies is not None else ''}" \
                 f"{_stoich if max_stoich else ''}"
        line_format = f"{{:<{max_fnm}}} | {{}} | {{}}" \
                      f"{' | {:^ 4}' if frequencies is not None else '{}'}" \
                      f"{f' | {{:<{max_stoich}}}' if max_stoich else '{}'}\n"
        # fname = 'distribution_overview.txt'
        with open(file, 'w') as file_:
            file_.write(header + '\n')
            names_line = ' ' * max_fnm + ' | ' + population_subheader + \
                         ' | ' + energies_subheader + \
                         (' |     ' if frequencies is not None else '') + \
                         (' | ' if max_stoich else '') + '\n'
            file_.write(names_line)
            file_.write('-' * len(header) + '\n')
            rows = zip_longest(
                filenames, values, popul, imaginary, stoichiometry, fillvalue=''
            )
            for fnm, vals, pops, imag, stoich in rows:
                p_line = '  '.join(
                    [f'{p:>{w}.4f}' for p, w in zip(pops, population_widths)]
                )
                v_line = '  '.join(
                    [f'{v:> {w}.{p}f}' for v, w, p
                     in zip(vals, energies_widths, precisions)]
                )
                line = line_format.format(fnm, p_line, v_line, imag, stoich)
                file_.write(line)
        logger.info('Energies collective export to text file done.')

    def bars(self, dest, band, bars, interfix=''):
        """Writes Bars objects to txt files (one for each conformer).

        Notes
        -----
        Filenames are generated in form of conformer_name[.interfix].txt

        Parameters
        ----------
        dest: string
            path to destination directory
        band: glassware.Bars
            object containing information about band at which transitions occur;
            it should be frequencies for vibrational data and wavelengths or
            excitation energies for electronic data
        bars: list of glassware.Bars
            Bars objects that are to be serialized; all should contain
            information for the same conformers
        interfix: string, optional
            string included in produced filenames, nothing is added if omitted
        """
        bars = [band] + bars
        genres = [bar.genre for bar in bars]
        headers = [self._header[genre] for genre in genres]
        widths = [self._formatters[genre][4:-4] for genre in genres]
        formatted = [f'{h: <{w}}' for h, w in zip(headers, widths)]
        values = zip(*[bar.values for bar in bars])
        for fname, values_ in zip(bars[0].filenames, values):
            filename = f"{'.'.join(fname.split('.')[:-1])}" \
                       f"{'.' if interfix else ''}{interfix}.txt"
            with open(os.path.join(dest, filename), 'w') as file:
                file.write('\t'.join(formatted))
                file.write('\n')
                for vals in zip(*values_):
                    line = '\t'.join(self._formatters[g].format(v)
                                     for v, g in zip(vals, genres))
                    file.write(line + '\n')
        logger.info('Bars export to text files done.')

    def spectra(self, dest, spectra, interfix=''):
        """Writes Spectra object to text files (one for each conformer).

        Notes
        -----
        Filenames are generated in form of conformer_name[.interfix].txt

        Parameters
        ----------
        dest: string
            path to destination directory
        spectra: glassware.Spectra
            Spectra object, that is to be serialized
        interfix: string, optional
            string included in produced filenames, nothing is added if omitted
        """
        abscissa = spectra.x
        title = f'{spectra.genre} calculated with peak width = {spectra.width}' \
                f' {spectra.units["width"]} and {spectra.fitting} ' \
                f'fitting, shown as {spectra.units["x"]} vs. ' \
                f'{spectra.units["y"]}'
        for fnm, values in zip(spectra.filenames, spectra.y):
            file_name = f"{'.'.join(fnm.split('.')[:-1])}" \
                        f"{'.' if interfix else ''}{interfix}.txt"
            file_path = os.path.join(dest, file_name)
            with open(file_path, 'w') as file:
                file.write(title + '\n')
                file.write(
                    '\n'.join(
                        f'{int(a):>4d}\t{v: .4f}'
                        for a, v in zip(abscissa, values)
                    )
                )
        logger.info('Spectra export to text files done.')

    def single_spectrum(self, file, spectrum, include_header=True):
        """Writes SingleSpectrum object to txt file.

        Parameters
        ----------
        file: string
            path to file
        spectrum: glassware.SingleSpectrum
            spectrum, that is to be serialized
        include_header: bool, optional
            determines if file should contain a header with metadata,
            True by default
        """
        title = f'{spectrum.genre} calculated with peak width = ' \
                f'{spectrum.width} {spectrum.units["width"]} and ' \
                f'{spectrum.fitting} fitting, shown as {spectrum.units["x"]} ' \
                f'vs. {spectrum.units["y"]}'
        with open(file, 'w') as file_:
            if include_header:
                file_.write(title + '\n')
                if spectrum.averaged_by:
                    file_.write(
                        f'{len(spectrum.filenames)} conformers averaged base on'
                        f' {self._header[spectrum.averaged_by]}\n'
                    )
            file_.write(
                '\n'.join(
                    # TO DO: probably should change when nmr introduced
                    f'{int(x):>4d}\t{y: .4f}' for x, y in
                    zip(spectrum.x, spectrum.y)
                )
            )
        logger.info('Spectrum export to text files done.')


class XlsxWriter(Writer, fmt='xlsx'):

    def write(self, dest, data):
        data = self.distribute_data(data)
        if data['energies']:
            file = os.path.join(dest, 'distribution.xlsx')
            self.energies(
                file, data['energies'], frequencies=data['frequencies'],
                stoichiometry=data['stoichiometry'],
                corrections=data['corrections'].values()
            )
        if data['vibra']:
            file = os.path.join(dest, 'bars.vibra.xlsx')
            self.bars(file, band=data['frequencies'], bars=data['vibra'])
        if data['electr']:
            file = os.path.join(dest, 'bars.electr.xlsx')
            self.bars(file, band=data['wavelengths'], bars=data['electr'])
        if data['other_bars']:
            # TO DO
            pass
        if data['spectra']:
            file = os.path.join(dest, 'spectra.xlsx')
            self.spectra(file, data['spectra'])
        if data['single']:
            file = os.path.join(dest, 'averaged_spectra.xlsx')
            self.single_spectrum(file, data['single'])
        if data['other']:
            # TO DO
            pass

    def energies(self, file, energies, frequencies=None,
                 stoichiometry=None, corrections=None):
        """Writes detailed information from multiple Energies objects to
         single xlsx file.

        Parameters
        ----------
        file: string
            path to file
        energies: list of glassware.Energies
            Energies objects that is to be exported
        frequencies: glassware.DataArray, optional
            DataArray object containing frequencies
        stoichiometry: glassware.InfoArray, optional
            InfoArray object containing stoichiometry information
        corrections: list of glassware.DataArray
            DataArray objects containing energies corrections"""
        wb = oxl.Workbook()
        ws = wb.active
        ens_no = len(energies)
        ws.title = 'Collective overview'
        headers = ['Gaussian output file', 'Populations / %',
                   'Energies / hartree']
        headers += ['Imag'] if frequencies is not None else []
        headers += ['Stoichiometry'] if stoichiometry is not None else []
        cells = ['A1', 'B1', f'{chr(66+ens_no)}1', f'{chr(66+2*ens_no)}1',
                 f'{chr(67+2*ens_no)}1']
        for header, cell in zip(headers, cells):
            ws[cell] = header
        names = [self._header[en.genre] for en in energies]
        ws.append([''] + names + names)
        ws.merge_cells('A1:A2')
        ws.merge_cells(f'B1:{chr(65+ens_no)}1')
        ws.merge_cells(f'{chr(66+ens_no)}1:{chr(65+2*ens_no)}1')
        if frequencies is not None or stoichiometry is not None:
            ws.merge_cells(f'{chr(66+2*ens_no)}1:{chr(66+2*ens_no)}2')
        if frequencies is not None and stoichiometry is not None:
            ws.merge_cells(f'{chr(67+2*ens_no)}1:{chr(67+2*ens_no)}2')
        ws.freeze_panes = 'A3'
        # data = self.ts.energies
        filenames = energies[0].filenames
        fmts = ['0'] + ['0.00%'] * len(energies) + \
               ['0.' + '0' * (8 if en.genre == 'scf' else 6) for en in
                energies] + \
               ['0', '0']
        values = [en.values for en in energies]
        populs = [en.populations for en in energies]
        imag = frequencies.imaginary if frequencies is not None else []
        stoich = stoichiometry.values if stoichiometry is not None else []
        rows = zip_longest(filenames, *populs, *values, imag, stoich)
        for row_num, values in enumerate(rows):
            filtered_values = ((f, v) for f, v in zip(fmts, values)
                               if v is not None)
            for col_num, (fmt, value) in enumerate(filtered_values):
                cell = ws.cell(row=row_num + 3, column=col_num + 1)
                cell.value = value
                cell.number_format = fmt
        # set cells width
        widths = [0] + [10] * ens_no + [16] * ens_no
        widths += [6] if frequencies is not None else []
        widths += [0] if stoichiometry is not None else []
        for column, width in zip(ws.columns, widths):
            if not width:
                width = max(len(str(cell.value)) for cell in column) + 2
            ws.column_dimensions[column[0].column].width = width
        # proceed to write detailed info on separate sheet for each energy
        corrs = {c.genre[:3]: c for c in corrections} \
            if corrections is not None else {}
        for en in energies:
            genre = en.genre
            corr = corrs.get(genre, None)
            fmts = ['0', '0.00%'] + ['0.0000'] * 2 + \
                   ['0.00000000' if genre == 'scf' else '0.000000'] * 2
            ws = wb.create_sheet(title=self._header[genre])
            ws.freeze_panes = 'A2'
            header = ['Gaussian output file', 'Population / %',
                      'Min. B. Factor', 'DE / (kcal/mol)',
                      'Energy / Hartree']
            header += ['Correction / Hartree'] if corr is not None else []
            ws.append(header)
            corr = corr.values if corr is not None else []
            rows = zip_longest(en.filenames, en.populations, en.min_factors,
                               en.deltas, en.values, corr)
            for row_num, values in enumerate(rows):
                filtered_values = ((f, v) for f, v in zip(fmts, values)
                                   if v is not None)
                for col_num, (fmt, value) in enumerate(filtered_values):
                    cell = ws.cell(row=row_num + 2, column=col_num + 1)
                    cell.value = value
                    cell.number_format = fmt
            # set cells width
            widths = [0, 15, 14, 15, 16, 19]
            for column, width in zip(ws.columns, widths):
                if not width:
                    width = max(len(str(cell.value)) for cell in column) + 2
                ws.column_dimensions[column[0].column].width = width
        wb.save(file)
        logger.info('Energies export to xlsx files done.')

    def bars(self, file, band, bars):
        """Writes Bars objects to xlsx file (one sheet for each conformer).

        Parameters
        ----------
        file: string
            path to file
        band: glassware.Bars
            object containing information about band at which transitions occur;
            it should be frequencies for vibrational data and wavelengths or
            excitation energies for electronic data
        bars: list of glassware.Bars
            Bars objects that are to be serialized; all should contain
            information for the same conformers"""
        wb = oxl.Workbook()
        wb.remove(wb.active)
        bars = [band] + bars
        genres = [bar.genre for bar in bars]
        headers = [self._header[genre] for genre in genres]
        widths = [max(len(h), 10) for h in headers]
        fmts = [self._excel_formats[genre] for genre in genres]
        values = list(zip(*[bar.values for bar in bars]))
        for fname, values_ in zip(bars[0].filenames, values):
            ws = wb.create_sheet(fname)
            ws.append(headers)
            ws.freeze_panes = 'B2'
            for column, width in zip(ws.columns, widths):
                ws.column_dimensions[column[0].column].width = width
            for col_num, (vals, fmt) in enumerate(zip(values_, fmts)):
                for row_num, v in enumerate(vals):
                    cell = ws.cell(row=row_num + 2, column=col_num + 1)
                    cell.value = v
                    cell.number_format = fmt
        wb.save(file)
        logger.info('Bars export to xlsx files done.')

    def spectra(self, file, spectra):
        wb = oxl.Workbook()
        del wb['Sheet']
        for spectra_ in spectra:
            ws = wb.create_sheet()
            ws.title = spectra_.genre
            ws.freeze_panes = 'B2'
            A0 = spectra_.units['x']
            ws.append([A0] + list(spectra_.filenames))
            title = f'{spectra_.genre} calculated with peak width = ' \
                    f'{spectra_.width} {spectra_.units["width"]} and ' \
                    f'{spectra_.fitting} fitting, shown as ' \
                    f'{spectra_.units["x"]} vs. {spectra_.units["y"]}'
            ws["A1"].comment = oxl.comments.Comment(title, 'Tesliper')
            for line in zip(spectra_.x, *spectra_.y):
                ws.append(line)
        wb.save(file)
        logger.info('Spectra export to xlsx file done.')

    def single_spectrum(self, file, spectra):
        # TO DO: add comment as in txt export
        # TO DO: think how to do it
        wb = oxl.Workbook()
        del wb['Sheet']
        for spc in spectra:
            ws = wb.create_sheet()
            ws.title = spc.genre + '_' + spc.averaged_by
            for row in zip(spc.x, spc.y):
                ws.append(row)
            wb.save(file)
        logger.info('Spectrum export to xlsx files done.')


class CsvWriter(Writer, fmt='csv'):

    def write(self, dest, data):
        data = self.distribute_data(data)
        if data['energies']:
            for en in data['energies']:
                file = os.path.join(dest, f'distribution.{en.genre}.csv')
                self.energies(
                    file, en, corrections=data['corrections'].get(en.genre)
                )
        if data['vibra']:
            self.bars(dest, band=data['frequencies'], bars=data['vibra'],
                      interfix='vibra')
        if data['electr']:
            self.bars(dest, band=data['wavelengths'], bars=data['electr'],
                      interfix='electr')
        if data['other_bars']:
            # TO DO
            pass
        if data['spectra']:
            for spc in data['spectra']:
                self.spectra(dest, spc, interfix=spc.genre)
        if data['single']:
            for spc in data['single']:
                interfix = f'.{spc.averaged_by}' if spc.averaged_by else ''
                file = os.path.join(dest, f'spectrum.{spc.genre+interfix}.csv')
                self.single_spectrum(file, spc)
        if data['other']:
            # TO DO
            pass


    def energies(self, file, energies, corrections=None,
                 include_header=True):
        """Writes Energies object to csv file.

        Parameters
        ----------
        file: string
            path to file
        energies: glassware.Energies
            Energies objects that is to be serialized
        corrections: glassware.DataArray, optional
            DataArray objects containing energies corrections
        include_header: bool, optional
            determines if file should contain a header with column names,
            True by default"""
        header = ['Gaussian output file']
        header += 'population min_factor delta energy'.split(' ')
        if corrections is not None:
            header += ['corrections']
            corr = corrections.values
        else:
            corr = []
        rows = zip_longest(energies.filenames, energies.populations,
                           energies.min_factors, energies.deltas,
                           energies.values, corr)
        with open(file, 'w', newline='') as file:
            csvwriter = csv.writer(file)
            if include_header:
                csvwriter.writerow(header)
            for row in rows:
                csvwriter.writerow(v for v in row if v is not None)
        logger.info('Energies export to csv files done.')

    def bars(self, dest, band, bars, include_header=True, interfix=''):
        """Writes Bars objects to csv files (one for each conformer).

        Notes
        -----
        Filenames are generated in form of {conformer_name}[.{interfix}].csv

        Parameters
        ----------
        dest: string
            path to destination directory
        band: glassware.Bars
            object containing information about band at which transitions occur;
            it should be frequencies for vibrational data and wavelengths or
            excitation energies for electronic data
        bars: list of glassware.Bars
            Bars objects that are to be serialized; all should contain
            information for the same conformers
        include_header: bool, optional
            determines if file should contain a header with column names,
            True by default,
        interfix: string, optional
            string included in produced filenames, nothing is added if omitted
        """
        bars = [band] + bars
        headers = [self._header[bar.genre] for bar in bars]
        values = zip(*[bar.values for bar in bars])
        for fname, values_ in zip(bars[0].filenames, values):
            filename = f"{'.'.join(fname.split('.')[:-1])}" \
                       f"{'.' if interfix else ''}{interfix}.csv"
            path = os.path.join(dest, filename)
            with open(path, 'w', newline='') as file:
                csvwriter = csv.writer(file)
                if include_header:
                    csvwriter.writerow(headers)
                for row in zip(*values_):
                    csvwriter.writerow(row)
        logger.info('Bars export to csv files done.')

    def spectra(self, dest, spectra, interfix='', include_header=True):
        abscissa = spectra.x
        for fnm, values in zip(spectra.filenames, spectra.y):
            file_name = f"{'.'.join(fnm.split('.')[:-1])}" \
                        f"{'.' if interfix else ''}{interfix}.csv"
            file_path = os.path.join(dest, file_name)
            with open(file_path, 'w', newline='') as file:
                csvwriter = csv.writer(file)
                if include_header:
                    # write header to file
                    pass
                for row in zip(abscissa, values):
                    csvwriter.writerow(row)
        logger.info('Spectra export to csv files done.')

    def single_spectrum(self, file, spectrum):
        with open(file, 'w', newline='') as file_:
            csvwriter = csv.writer(file_)
            for row in zip(spectrum.x, spectrum.y):
                csvwriter.writerow(row)
        logger.info('Spectrum export to csv files done.')
